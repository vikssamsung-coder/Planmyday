"""Admin daily analysis — per-user task completion, goal alignment, and effort, for one day.

Pure assembly from existing tables (tasks, day_goals, effort_kras) — no schema change. Used by
the Admin → Analysis tab: an in-app dashboard plus a downloadable Excel.
"""
import io

import classify
import nudge
import schemas
import storage


def _effort_kra_columns(uk):
    """The Effort-matrix KRA columns for a user (mirrors app._effort_kra_columns without an
    app import so there's no circular dependency): the user's saved list, else seeded from
    their monthly-target KPI names."""
    saved = storage.get_effort_kras(uk)
    if saved:
        return saved
    names = []
    try:
        df = storage._read(storage._targets_path(uk), schemas.MONTHLY_TARGETS)
        if not df.empty:
            for n in df["kpi_name"].tolist():
                n = str(n).strip()
                if n and n not in names:
                    names.append(n)
    except Exception:
        pass
    return names


def user_analysis(uk, user, date):
    """Compute one user's analysis for `date`: task completion, horizon split, goal alignment,
    targets vs achievement, and the effort matrix. Returns a dict."""
    tasks = storage.get_tasks(uk, date)
    goals = [g for g in storage.get_day_goals(uk, date) if g["heading"]]
    headings = [g["heading"] for g in goals]

    total = done = open_ = dropped = today_h = build_h = aligned = unaligned = 0
    if tasks is not None and len(tasks):
        for _, t in tasks.iterrows():
            status = str(t.get("status", "") or "")
            if status == "Dropped":
                dropped += 1
                continue
            total += 1
            if status == "Done":
                done += 1
            else:
                open_ += 1
            if str(t.get("horizon", "")) == "Build":
                build_h += 1
            else:
                today_h += 1
            dg = t.get("day_goal", "") or ""
            if dg and nudge.goal_served(dg, headings):
                aligned += 1
            else:
                unaligned += 1
    completion = round(100 * done / total) if total else 0

    targets_set = len(goals)
    targets_done = sum(1 for g in goals if str(g.get("achieved", "") or "").strip())

    kras = _effort_kra_columns(uk)
    rows, cols, counts, row_tot, col_tot, grand = classify.build_matrix(tasks, kras)
    unaligned_effort = col_tot.get(classify.UNASSIGNED, 0)
    top_kra = ""
    if grand:
        cand = max(cols, key=lambda c: col_tot.get(c, 0))
        if col_tot.get(cand, 0) > 0 and cand != classify.UNASSIGNED:
            top_kra = cand

    return {
        "user_key": uk,
        "name": user.get("name", uk),
        "role": str(user.get("role", "")).replace("_", " ").title(),
        "closed": storage.is_day_closed(uk, date),
        "tasks_total": total, "done": done, "open": open_, "dropped": dropped,
        "completion": completion, "today_h": today_h, "build_h": build_h,
        "aligned": aligned, "unaligned": unaligned,
        "targets_set": targets_set, "targets_done": targets_done, "goals": goals,
        "effort_grand": grand, "top_kra": top_kra, "unaligned_effort": unaligned_effort,
        "matrix": (rows, cols, counts, row_tot, col_tot, grand),
        "tasks": tasks,
    }


def build_xlsx(date, analyses):
    """A workbook: a Summary sheet (one row per user) + a detail sheet per user (targets,
    tasks, effort matrix). Returns xlsx bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append([f"Team daily analysis — {date}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    headers = ["User", "Role", "Tasks", "Done", "Open", "Dropped", "Completion %",
               "Today", "Build", "Aligned", "Unaligned", "Targets set", "Targets done",
               "Effort tasks", "Top KRA", "Unaligned effort", "Closed day"]
    ws.append(headers)
    hdr_row = ws.max_row
    for a in analyses:
        ws.append([a["name"], a["role"], a["tasks_total"], a["done"], a["open"], a["dropped"],
                   a["completion"], a["today_h"], a["build_h"], a["aligned"], a["unaligned"],
                   a["targets_set"], a["targets_done"], a["effort_grand"], a["top_kra"] or "—",
                   a["unaligned_effort"], "Yes" if a["closed"] else "No"])
    for c in ws[hdr_row]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2D4A5E")
        c.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        w = max((len(str(cell.value)) for cell in col if cell.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 8), 40)

    def _sheet_title(a):
        base = (a["name"] or a["user_key"] or "user")
        bad = '[]:*?/\\'
        base = "".join(ch for ch in base if ch not in bad)[:26] or "user"
        title = base
        i = 2
        while title in wb.sheetnames:
            title = f"{base[:24]}{i}"
            i += 1
        return title

    for a in analyses:
        wsd = wb.create_sheet(title=_sheet_title(a))
        wsd.append([f"{a['name']} — {date}"])
        wsd["A1"].font = Font(bold=True, size=13)
        wsd.append([])
        wsd.append(["Targets vs achievement"])
        wsd[wsd.max_row][0].font = Font(bold=True)
        wsd.append(["Heading", "Target", "Achieved"])
        for g in a["goals"]:
            wsd.append([g["heading"], g.get("target_number", ""), g.get("achieved", "")])
        if not a["goals"]:
            wsd.append(["(no targets set)", "", ""])
        wsd.append([])
        wsd.append(["Tasks"])
        wsd[wsd.max_row][0].font = Font(bold=True)
        wsd.append(["Task", "Goal", "Horizon", "Status"])
        tdf = a["tasks"]
        if tdf is not None and len(tdf):
            for _, t in tdf.iterrows():
                wsd.append([t.get("title", ""), t.get("day_goal", "") or "—",
                            t.get("horizon", "") or "Today", t.get("status", "")])
        else:
            wsd.append(["(no tasks)", "", "", ""])
        wsd.append([])
        wsd.append(["Effort matrix (task counts)"])
        wsd[wsd.max_row][0].font = Font(bold=True)
        rows, cols, counts, row_tot, col_tot, grand = a["matrix"]
        wsd.append([""] + list(cols) + ["TOTAL"])
        for r in rows:
            wsd.append([r] + [counts[r][c] for c in cols] + [row_tot[r]])
        wsd.append(["TOTAL"] + [col_tot[c] for c in cols] + [grand])
        for col in wsd.columns:
            w = max((len(str(cell.value)) for cell in col if cell.value is not None), default=8)
            wsd.column_dimensions[col[0].column_letter].width = min(max(w + 2, 8), 44)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
